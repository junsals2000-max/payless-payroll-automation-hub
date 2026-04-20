#!/usr/bin/env python3
"""
Bank Statement Categorizer — Payless Automation Hub
====================================================
Usage:
    python3 categorize.py --bank bank.xlsx --output out.xlsx
                          [--wise wise.pdf] [--wu wu_history.xlsx]
                          [--empmap employee_map.json]

Inputs:
  --bank   (required)  Raw bank statement XLSX (Date, Description, Amount, Running Bal.)
  --wise   (optional)  Wise transaction history PDF (recipient name + USD amounts)
  --wu     (optional)  Western Union transaction history XLSX (MTCN, Receiving details)
  --empmap (optional)  Path to employee_map.json (defaults to bank_categorizer_config.json in data/)

Output:
  Categorized Excel with SUM formulas.
  Uncategorized rows highlighted yellow at the bottom.
"""

import sys
import os
import re
import json
import argparse
from pathlib import Path
from collections import defaultdict, deque

# ──────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ──────────────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent.resolve()

# Default employee map: data/bank_categorizer_config.json (next to the hub)
DEFAULT_EMP_MAP = SCRIPT_DIR.parent.parent / "data" / "bank_categorizer_config.json"

# Keywords that identify vendor/overhead transactions by description
MATERIAL_KEYWORDS = ["ETERNITY FLOORING", "M S INTERNATIONA", "NEW AGE SURFACES", "DECOREATIVE"]
OVERHEAD_KEYWORDS = ["CLAUDE.AI SUBSCRIPTION", "SHELL OIL"]

# Transaction block order (how rows are grouped in the output)
GROUP_ORDER = [
    "MATERIAL",
    "OVERHEAD",
    "ADV_ZELLE",
    "PC_PAYROLL",
    "EVENT",
    "RECRUITER",
    "OFFICE_PAYROLL",
    "ADV_MARKETING",
    "REMOTE_SALES",
    "SC_PAYROLL",
    "BRAND_AMBASSADOR",
]

# Summary block order (rows 1-11) — label shown in column C, SUM formula in column D
SUMMARY_ROWS = [
    ("MATERIAL",          "MATERIAL "),
    ("ADV_ZELLE",         "ADVERTISING "),
    ("ADV_MARKETING",     "ADVERTISING PAYROLL"),
    ("RECRUITER",         "RECRUITER"),
    ("OVERHEAD",          "OVERHEAD "),
    ("PC_PAYROLL",        "PC PAYROLL "),
    ("SC_PAYROLL",        "SC PAYROLL"),
    ("BRAND_AMBASSADOR",  "BRAND AMBASSADOR PAYROLL "),
    ("OFFICE_PAYROLL",    "OFFICE PAYROLL "),
    ("REMOTE_SALES",      "REMOTE SALES"),
    ("EVENT",             "EVENT"),
]

PAYPAL_LABEL = "Office Payroll - Elcia"
PAYPAL_GROUP = "OFFICE_PAYROLL"

ZELLE_BOXTRUCK_LABEL = "ADVERTISING-EVENT - PER TAMAR The box truck goes under events"
ZELLE_BOXTRUCK_GROUP = "ADV_ZELLE"

# ── Color map: group -> (fill_hex, font_hex)  (no "#", 6-char RRGGBB) ─────────
GROUP_COLORS = {
    "MATERIAL":          ("F2F2F2", "000000"),
    "OVERHEAD":          ("FFE699", "000000"),
    "ADV_ZELLE":         ("FFC000", "000000"),
    "PC_PAYROLL":        ("9DC3E6", "000000"),
    "EVENT":             ("1F3864", "FFFFFF"),
    "RECRUITER":         ("375623", "FFFFFF"),
    "OFFICE_PAYROLL":    ("92D050", "000000"),
    "ADV_MARKETING":     ("FFC000", "000000"),
    "REMOTE_SALES":      ("808080", "FFFFFF"),
    "SC_PAYROLL":        ("F4B942", "000000"),
    "BRAND_AMBASSADOR":  ("FCE4D6", "000000"),
}

# ──────────────────────────────────────────────────────────────────────────────
# PARSE WISE PDF
# ──────────────────────────────────────────────────────────────────────────────

def _clean_wise_name(raw):
    raw = re.sub(r"\s+[\d,]+(?:\.\d+)?\s+[A-Z]{2,5}.*$", "", raw)
    raw = re.sub(r"(\s+[a-z]{1,4}\.?\s*)+$", "", raw)
    return raw.strip()


def parse_wise_pdf(pdf_path):
    import pdfplumber

    result = defaultdict(deque)

    with pdfplumber.open(pdf_path) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    USD_RE = re.compile(r"([\d,]+(?:\.\d{1,2})?)\s+USD")

    lines = full_text.split("\n")
    n = len(lines)
    i = 0
    while i < n:
        line = lines[i].strip()

        sent_match = re.match(r"^Sent\s*\|\s*(.+)", line)
        if sent_match:
            raw_name = _clean_wise_name(sent_match.group(1))

            all_usd = []
            for j in range(i + 1, min(i + 16, n)):
                next_line = lines[j].strip()
                if re.match(r"^Sent\s*\|", next_line):
                    break
                found = USD_RE.findall(next_line)
                all_usd.extend(found)

            if all_usd:
                total = float(all_usd[-1].replace(",", ""))
                result[total].append(raw_name)

        i += 1

    return result


# ──────────────────────────────────────────────────────────────────────────────
# PARSE WU XLSX
# ──────────────────────────────────────────────────────────────────────────────

def parse_wu_xlsx(xlsx_path):
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    result = {}
    mtcn_col = recv_col = None

    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue

        row_str = [str(c).strip() if c else "" for c in row]

        if "MTCN" in row_str:
            mtcn_col = row_str.index("MTCN")
            recv_col = next(
                (i for i, h in enumerate(row_str) if "receiving" in h.lower()), None
            )
            continue

        if mtcn_col is None or recv_col is None:
            continue

        mtcn = row_str[mtcn_col]
        recv = row_str[recv_col]

        if not mtcn or not recv:
            continue

        digits = re.sub(r"\D", "", mtcn)
        if len(digits) < 4:
            continue
        last4 = digits[-4:]

        name_match = re.match(
            r"^(.+?)\s+(?:Bank account|Cash at agent|Mobile wallet)", recv, re.IGNORECASE
        )
        name = name_match.group(1).strip() if name_match else recv.split()[0]
        result[last4] = name

    return result


# ──────────────────────────────────────────────────────────────────────────────
# LOAD EMPLOYEE MAP
# ──────────────────────────────────────────────────────────────────────────────

def load_employee_map(path=None):
    if path is None:
        path = DEFAULT_EMP_MAP
    with open(path, encoding="utf-8") as f:
        raw = json.load(f)
    # Drop comment keys and metadata keys (starting with _)
    return {k: v for k, v in raw.items() if not k.startswith("_")}


def lookup_employee(name, emp_map):
    key = name.lower().strip()

    if key in emp_map:
        e = emp_map[key]
        return e["label"], e["group"]

    for map_key, entry in emp_map.items():
        if map_key in key or key in map_key:
            return entry["label"], entry["group"]

    return None, None


def _suggest_label(name):
    """Title-case the name as a suggested label."""
    return " ".join(w.capitalize() for w in name.split())


def _suggest_group(name, emp_map):
    """
    Try to guess the most likely group from partial name matches,
    or return None if no clue.
    """
    key = name.lower().strip()
    for map_key, entry in emp_map.items():
        words = key.split()
        for w in words:
            if len(w) > 3 and w in map_key:
                return entry["group"]
    return None


# ──────────────────────────────────────────────────────────────────────────────
# CATEGORIZE ONE TRANSACTION
# ──────────────────────────────────────────────────────────────────────────────

def categorize_one(date, desc, amount, wise_amounts, wu_map, emp_map, paypal_count, unmatched_out):
    """
    Returns (label, group).
    unmatched_out is a list that collects unmatched employee suggestions.
    """
    if not desc:
        return None, "UNCATEGORIZED"

    desc_u = desc.upper()

    # ── Material vendors ──────────────────────────────────────────────────────
    for kw in MATERIAL_KEYWORDS:
        if kw in desc_u:
            return "MATERIAL", "MATERIAL"

    # ── Overhead vendors ──────────────────────────────────────────────────────
    for kw in OVERHEAD_KEYWORDS:
        if kw in desc_u:
            return "OVERHEAD", "OVERHEAD"

    # ── Zelle box truck ───────────────────────────────────────────────────────
    if "ZELLE" in desc_u and ("ONE WAY TOWING" in desc_u or "BOX TRUCK" in desc_u):
        return ZELLE_BOXTRUCK_LABEL, ZELLE_BOXTRUCK_GROUP

    # ── Wise ──────────────────────────────────────────────────────────────────
    if "WISE INC" in desc_u or "DES:WISE" in desc_u:
        amt = abs(amount) if amount else None
        if amt and amt in wise_amounts and len(wise_amounts[amt]) > 0:
            name = wise_amounts[amt].popleft()
            label, group = lookup_employee(name, emp_map)
            if label:
                return label, group
            else:
                suggested_label = _suggest_label(name)
                suggested_group = _suggest_group(name, emp_map)
                print(f"UNMATCHED_SUGGESTION: source=Wise | raw_name={name} | suggested_label={suggested_label} | suggested_group={suggested_group or ''}")
                unmatched_out.append({
                    "source": "Wise",
                    "raw_name": name,
                    "suggested_label": suggested_label,
                    "suggested_group": suggested_group or "",
                    "amount": amt,
                })
        elif amt:
            print(f"  ⚠️  Wise: no PDF match for amount ${amt:.2f} — marked UNCATEGORIZED")
        return None, "UNCATEGORIZED"

    # ── Western Union ─────────────────────────────────────────────────────────
    if "WESTERN UNION" in desc_u:
        cap_match = re.search(r"ID:X+(\d{4,6})", desc)
        if cap_match:
            last4 = cap_match.group(1)[-4:]
            wu_name = wu_map.get(last4)
            if wu_name:
                label, group = lookup_employee(wu_name, emp_map)
                if label:
                    return label, group
                suggested_label = _suggest_label(wu_name)
                suggested_group = _suggest_group(wu_name, emp_map)
                print(f"UNMATCHED_SUGGESTION: source=Western Union | raw_name={wu_name} | suggested_label={suggested_label} | suggested_group={suggested_group or ''}")
                unmatched_out.append({
                    "source": "Western Union",
                    "raw_name": wu_name,
                    "suggested_label": suggested_label,
                    "suggested_group": suggested_group or "",
                    "amount": None,
                })
            else:
                print(f"  ⚠️  WU: MTCN last4 '{last4}' not found in WU xlsx — marked UNCATEGORIZED")
        return None, "UNCATEGORIZED"

    # ── PayPal ────────────────────────────────────────────────────────────────
    if "PAYPAL" in desc_u:
        paypal_count[0] += 1
        if paypal_count[0] > 1:
            print(f"  ⚠️  MULTIPLE PAYPAL TRANSACTIONS DETECTED (#{paypal_count[0]}): {desc}")
        return PAYPAL_LABEL, PAYPAL_GROUP

    # ── Payoneer ──────────────────────────────────────────────────────────────
    if "PAYONEER" in desc_u:
        pay_to = re.search(r"Pay To:\s*(.+?)(?:\s+Payment Date|$)", desc, re.IGNORECASE)
        if pay_to:
            name = pay_to.group(1).strip()
            label, group = lookup_employee(name, emp_map)
            if label:
                return label, group
            suggested_label = _suggest_label(name)
            suggested_group = _suggest_group(name, emp_map)
            print(f"UNMATCHED_SUGGESTION: source=Payoneer | raw_name={name} | suggested_label={suggested_label} | suggested_group={suggested_group or ''}")
            unmatched_out.append({
                "source": "Payoneer",
                "raw_name": name,
                "suggested_label": suggested_label,
                "suggested_group": suggested_group or "",
                "amount": None,
            })
        return None, "UNCATEGORIZED"

    # ── Direct bank transfer with named recipient ─────────────────────────────
    if "TRANSFER CARPET WAGON" in desc_u:
        name_match = re.search(
            r"TRANSFER CARPET WAGON GLENDAL:(.+?)\s+Confirmation", desc, re.IGNORECASE
        )
        if name_match:
            name = name_match.group(1).strip()
            label, group = lookup_employee(name, emp_map)
            if label:
                return label, group
            suggested_label = _suggest_label(name)
            suggested_group = _suggest_group(name, emp_map)
            print(f"UNMATCHED_SUGGESTION: source=Bank Transfer | raw_name={name} | suggested_label={suggested_label} | suggested_group={suggested_group or ''}")
            unmatched_out.append({
                "source": "Bank Transfer",
                "raw_name": name,
                "suggested_label": suggested_label,
                "suggested_group": suggested_group or "",
                "amount": None,
            })
        return None, "UNCATEGORIZED"

    return None, "UNCATEGORIZED"


# ──────────────────────────────────────────────────────────────────────────────
# PARSE BANK STATEMENT
# ──────────────────────────────────────────────────────────────────────────────

def parse_bank_xlsx(xlsx_path):
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    transactions = []
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        cell0 = str(row[0]).strip() if row[0] else ""
        cell1 = str(row[1]).strip() if row[1] else ""

        if cell0 == "Date" and cell1 == "Description":
            header_found = True
            continue

        if not header_found:
            continue

        date   = row[0]
        desc   = str(row[1]).strip() if row[1] else None
        amount = row[2]

        if desc and "BEGINNING BALANCE" in desc.upper():
            continue

        if not date or not hasattr(date, "strftime"):
            continue

        transactions.append((date, desc, amount))

    return transactions


# ──────────────────────────────────────────────────────────────────────────────
# BUILD OUTPUT XLSX
# ──────────────────────────────────────────────────────────────────────────────

def _make_fill(hex6):
    from openpyxl.styles import PatternFill
    return PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")

def _make_font(hex6, bold=False):
    from openpyxl.styles import Font
    return Font(color=hex6, bold=bold)

def _apply_row_style(ws, row_num, fill, font, num_cols=4):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font


def build_output_xlsx(grouped, output_path):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "in"

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 82
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 16

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font   = Font(bold=True)

    group_styles = {}
    for grp, (fill_hex, font_hex) in GROUP_COLORS.items():
        group_styles[grp] = (
            _make_fill(fill_hex),
            _make_font(font_hex),
        )

    # ── Summary block labels (rows 1-11, column C) ────────────────────────────
    for idx, (_, label) in enumerate(SUMMARY_ROWS, start=1):
        ws.cell(row=idx, column=3, value=label)

    # ── Header row 12 ─────────────────────────────────────────────────────────
    for col, text in enumerate(["Date", "Description", "Amount"], start=1):
        cell = ws.cell(row=12, column=col, value=text)
        cell.font = bold_font

    # ── Transaction block (starts at row 13) ──────────────────────────────────
    FIRST_DATA_ROW = 13
    current_row = FIRST_DATA_ROW
    group_ranges = {}

    for group_key in GROUP_ORDER:
        txns = grouped.get(group_key, [])
        if not txns:
            continue

        fill, font = group_styles.get(group_key, (_make_fill("FFFFFF"), _make_font("000000")))

        start_row = current_row
        for date, desc, amount, label in txns:
            _apply_row_style(ws, current_row, fill, font)
            ws.cell(row=current_row, column=1, value=date)
            ws.cell(row=current_row, column=1).number_format = "M/D/YYYY"
            ws.cell(row=current_row, column=2, value=desc)
            if amount is not None:
                ws.cell(row=current_row, column=3, value=amount)
                ws.cell(row=current_row, column=3).number_format = "#,##0.00"
            if label:
                ws.cell(row=current_row, column=4, value=label)
            current_row += 1

        group_ranges[group_key] = (start_row, current_row - 1)

    last_categorized_row = current_row - 1

    # ── Uncategorized rows (yellow, no label) ────────────────────────────────
    for date, desc, amount, _ in grouped.get("UNCATEGORIZED", []):
        for col in range(1, 5):
            ws.cell(row=current_row, column=col).fill = yellow_fill
        ws.cell(row=current_row, column=1, value=date)
        ws.cell(row=current_row, column=1).number_format = "M/D/YYYY"
        ws.cell(row=current_row, column=2, value=desc)
        if amount is not None:
            ws.cell(row=current_row, column=3, value=amount)
            ws.cell(row=current_row, column=3).number_format = "#,##0.00"
        current_row += 1

    # ── Back-fill summary formulas (column D, rows 1-11) ─────────────────────
    for row_idx, (group_key, _) in enumerate(SUMMARY_ROWS, start=1):
        if group_key in group_ranges:
            r1, r2 = group_ranges[group_key]
            formula = f"=SUM(C{r1})" if r1 == r2 else f"=SUM(C{r1}:C{r2})"
        else:
            formula = "=0"
        ws.cell(row=row_idx, column=4, value=formula)

    # ── Row 12 totals ─────────────────────────────────────────────────────────
    ws.cell(row=12, column=4, value="=SUM(D1:D11)").font = bold_font
    if last_categorized_row >= FIRST_DATA_ROW:
        ws.cell(
            row=12, column=5,
            value=f"=SUM(C{FIRST_DATA_ROW}:C{last_categorized_row})"
        ).font = bold_font

    wb.save(output_path)
    return output_path


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Bank Statement Categorizer")
    parser.add_argument("--bank",   required=True,  help="Bank statement XLSX (required)")
    parser.add_argument("--wise",   default=None,   help="Wise PDF (optional)")
    parser.add_argument("--wu",     default=None,   help="Western Union XLSX (optional)")
    parser.add_argument("--empmap", default=None,   help="Path to employee_map JSON (optional)")
    parser.add_argument("--output", default=None,   help="Output XLSX path")
    args = parser.parse_args()

    bank_path   = args.bank
    wise_path   = args.wise
    wu_path     = args.wu
    emp_map_path = args.empmap
    output_path = args.output or str(Path(bank_path).with_stem(Path(bank_path).stem + "_categorized"))

    print("📂 Loading employee map...")
    emp_map = load_employee_map(emp_map_path)
    print(f"   {len(emp_map)} employees loaded")

    wise_amounts = defaultdict(deque)
    if wise_path:
        print("📄 Parsing Wise PDF...")
        wise_amounts = parse_wise_pdf(wise_path)
        print(f"   {sum(len(v) for v in wise_amounts.values())} Wise recipients found")
    else:
        print("📄 Wise PDF not provided — skipping")

    wu_map = {}
    if wu_path:
        print("💳 Parsing WU history...")
        wu_map = parse_wu_xlsx(wu_path)
        print(f"   {len(wu_map)} WU recipients found")
    else:
        print("💳 WU history not provided — skipping")

    print("📋 Parsing bank statement...")
    transactions = parse_bank_xlsx(bank_path)
    print(f"   {len(transactions)} transactions found")

    print("🏷️  Categorizing transactions...")
    paypal_count  = [0]
    unmatched_out = []

    grouped = defaultdict(list)

    for date, desc, amount in transactions:
        label, group = categorize_one(
            date, desc, amount, wise_amounts, wu_map, emp_map, paypal_count, unmatched_out
        )
        grouped[group].append((date, desc, amount, label))

    # Warn about any Wise amounts left unconsumed
    leftover = {amt: list(names) for amt, names in wise_amounts.items() if names}
    if leftover:
        print("\n  ⚠️  Wise PDF entries NOT matched to any bank transaction:")
        for amt, names in leftover.items():
            for n in names:
                print(f"       ${amt:.2f}  — {n}")

    if paypal_count[0] > 1:
        print(f"\n  🚨 ACTION REQUIRED: {paypal_count[0]} PayPal transactions found.")

    # Category counts
    print("\n📊 Categorization summary:")
    total_categorized = 0
    for g in GROUP_ORDER:
        count = len(grouped.get(g, []))
        if count:
            print(f"   {g:<20} {count:>3} transactions")
            total_categorized += count
    uncategorized = len(grouped.get("UNCATEGORIZED", []))
    print(f"   {'UNCATEGORIZED':<20} {uncategorized:>3} transactions (yellow)")
    print(f"   {'TOTAL':<20} {total_categorized + uncategorized:>3}")

    print(f"\n📁 Building output Excel → {output_path}")
    build_output_xlsx(grouped, output_path)
    print("✅  Done!")

    return output_path


if __name__ == "__main__":
    main()
