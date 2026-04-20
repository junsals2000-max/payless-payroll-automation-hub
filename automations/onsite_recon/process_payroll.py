#!/usr/bin/env python3
"""
Payroll Report Generator
Transforms a Gusto payroll file (CSV, XLS, or XLSX) into a formatted Excel report.

Usage:
    python3 process_payroll.py --input <file_path> --output <xlsx_path> --depts <depts_json>

Arguments:
    --input   Path to the Gusto payroll file (.csv, .xls, or .xlsx)
    --output  Path for the output Excel file
    --depts   Path to departments.json (default: references/departments.json next to this script)
"""

import argparse
import csv
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Constants ─────────────────────────────────────────────────────────────────

DEPT_ORDER = [
    'INSTALLERS', 'PC-WAREHOUSE', 'SALES TEAM', 'OFFICE',
    'PC-OFFICE', 'CALL CENTER', 'EVENT TEAM',
]
DEPT_RANK = {d: i + 1 for i, d in enumerate(DEPT_ORDER)}
DEPT_RANK['No Group'] = 8

DEPT_COLORS = {
    'INSTALLERS':   'FFF200',
    'PC-WAREHOUSE': 'A9D08E',
    'SALES TEAM':   'F4B6C2',
    'OFFICE':       'D9B2DE',
    'PC-OFFICE':    '9BC2E6',
    'CALL CENTER':  '8FD3F4',
    'EVENT TEAM':   'F8CBAD',
    'No Group':     'FFFFFF',
}

GREY   = 'D9D9D9'
FONT   = 'Arial'

# Column indices (0-based) — Gusto Payroll Details export layout
C_NAME      = 0
C_DATE      = 1
C_PERIOD    = 2
C_GROSS     = 10
C_NET       = 38
C_EMP_TAX   = 40
C_TOTAL     = 50


# ── Name normalization ─────────────────────────────────────────────────────────

SUFFIX_RE = re.compile(r'\b(JR|SR|III|II|IV|JUNIOR|SENIOR)\.?\b', re.IGNORECASE)

def normalize_name(name: str) -> str:
    """Normalize to LASTNAME,FIRSTNAME (uppercase, no middle initials/suffixes)."""
    name = re.sub(r'[*.]', '', name)
    name = re.sub(r'\s+', ' ', name).strip().upper()
    name = SUFFIX_RE.sub('', name)
    name = re.sub(r'\s+', ' ', name).strip()

    if ',' in name:
        last, rest = name.split(',', 1)
        parts = rest.strip().split()
        first = parts[0] if parts else ''
        return f"{last.strip()},{first}"

    # No comma — shouldn't happen with well-formed Gusto exports, but handle anyway
    parts = name.split()
    if len(parts) >= 2:
        # Assume "Lastname Firstname" or "Firstname Lastname"
        # Return as-is normalized; matching will still attempt lookup
        return name
    return name


def build_lookup(depts_path: str):
    """Build a normalized-name → department lookup dict."""
    with open(depts_path) as f:
        data = json.load(f)

    lookup = {}
    for dept, names in data.get('departments', {}).items():
        for name in names:
            key = normalize_name(name)
            lookup[key] = dept

    # Also index by nickname so both legal and nickname resolve to the same dept
    for legal_norm, nick_norm in data.get('nickname_overrides', {}).items():
        if legal_norm in lookup:
            lookup[nick_norm] = lookup[legal_norm]
        elif nick_norm in lookup:
            lookup[legal_norm] = lookup[nick_norm]

    return lookup


# ── Money parsing ──────────────────────────────────────────────────────────────

def parse_money(val: str) -> float:
    if not val or not val.strip():
        return 0.0
    val = re.sub(r'[\$,\s]', '', val)
    try:
        return float(val)
    except ValueError:
        return 0.0


# ── File reading (CSV / XLS / XLSX) ───────────────────────────────────────────

def _read_rows_from_excel(path: str):
    """Read all rows from an XLS or XLSX file as lists of strings."""
    ext = Path(path).suffix.lower()
    rows = []

    if ext == '.xlsx':
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows.append(['' if v is None else str(v).strip() for v in row])

    elif ext == '.xls':
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        for r in range(ws.nrows):
            rows.append([str(ws.cell_value(r, c)).strip() for c in range(ws.ncols)])

    else:
        raise ValueError(f"Unsupported Excel format: {ext}")

    return rows


def read_payroll_file(path: str):
    """Read payroll file (CSV, XLS, XLSX), return list of employee dicts."""
    ext = Path(path).suffix.lower()

    if ext == '.csv':
        with open(path, newline='', encoding='utf-8-sig') as f:
            all_rows = list(csv.reader(f))
    else:
        all_rows = _read_rows_from_excel(path)

    # Find the real header row (first row whose col-A value is exactly "Name")
    header_idx = next(
        (i for i, r in enumerate(all_rows) if r and r[0].strip() == 'Name'),
        None
    )
    if header_idx is None:
        raise ValueError("Could not find 'Name' header row in the file.")

    data_rows = all_rows[header_idx + 1:]
    employees = []
    for orig_idx, row in enumerate(data_rows):
        if not row or not row[0].strip():
            continue
        name = row[0].strip()
        if name.lower() == 'total':
            continue

        def safe(col, r=row):
            return r[col] if len(r) > col else ''

        employees.append({
            'name':      name,
            'pay_date':  safe(C_DATE),
            'period':    safe(C_PERIOD),
            'gross':     parse_money(safe(C_GROSS)),
            'net':       parse_money(safe(C_NET)),
            'emp_tax':   parse_money(safe(C_EMP_TAX)),
            'total':     parse_money(safe(C_TOTAL)),
            'norm':      normalize_name(name),
            'orig_idx':  orig_idx,
        })

    return employees


# ── Department assignment & sorting ───────────────────────────────────────────

def assign_and_sort(employees, lookup):
    no_group = []
    for emp in employees:
        dept = lookup.get(emp['norm'], 'No Group')
        emp['dept'] = dept
        emp['rank'] = DEPT_RANK.get(dept, 8)
        if dept == 'No Group':
            no_group.append(emp['name'])

    # Stable sort: rank first, preserve original order within each dept
    employees.sort(key=lambda e: (e['rank'], e['orig_idx']))
    return employees, no_group


# ── Excel helpers ──────────────────────────────────────────────────────────────

def cell_fill(color):
    return PatternFill('solid', start_color=color, end_color=color)

def write_cell(ws, row, col, value, fill, font, fmt=None, align='left'):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill
    c.font = font
    if fmt:
        c.number_format = fmt
    c.alignment = Alignment(horizontal=align, vertical='center')
    return c


# ── Main Excel builder ─────────────────────────────────────────────────────────

def build_excel(employees, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Payroll Report'

    dark_fill = cell_fill('595959')
    grey_fill = cell_fill(GREY)
    money_fmt = '$#,##0.00'

    # ── Row 1: Column headers ───────────────────────────────────────
    hdr_cols = [
        'Name', 'Pay Date', 'Time Period',
        'Gross Pay – Total', 'Net Pay',
        'Employer Taxes & Contributions – Total', 'Total Payroll Cost',
    ]
    hdr_font = Font(name=FONT, bold=True, size=10, color='FFFFFF')
    for c, h in enumerate(hdr_cols, 1):
        write_cell(ws, 1, c, h, dark_fill, hdr_font, align='center')
    ws.row_dimensions[1].height = 30

    # ── Rows 2+: Employee data ──────────────────────────────────────
    curr = 2
    dept_row_ranges = {}   # dept → (first_row, last_row) for SUM formulas

    for emp in employees:
        dept  = emp['dept']
        color = DEPT_COLORS.get(dept, 'FFFFFF')
        fill  = cell_fill(color)
        font  = Font(name=FONT, size=10)

        if dept not in dept_row_ranges:
            dept_row_ranges[dept] = [curr, curr]
        else:
            dept_row_ranges[dept][1] = curr

        values = [
            emp['name'], emp['pay_date'], emp['period'],
            emp['gross'], emp['net'], emp['emp_tax'], emp['total'],
        ]
        for c, v in enumerate(values, 1):
            align = 'left' if c <= 3 else 'right'
            fmt   = money_fmt if c >= 4 else None
            write_cell(ws, curr, c, v, fill, font, fmt=fmt, align=align)

        curr += 1

    last_data_row = curr - 1

    # ── Grand Total row ─────────────────────────────────────────────
    gt_row  = curr
    gt_fill = grey_fill
    gt_font = Font(name=FONT, bold=True, size=10)

    write_cell(ws, gt_row, 1, 'Grand Total', gt_fill, gt_font, align='left')
    write_cell(ws, gt_row, 2, None,          gt_fill, gt_font)
    write_cell(ws, gt_row, 3, None,          gt_fill, gt_font)

    for c, col in enumerate(['D', 'E', 'F', 'G'], 4):
        write_cell(ws, gt_row, c,
                   f'=SUM({col}2:{col}{last_data_row})',
                   gt_fill, gt_font, fmt=money_fmt, align='right')

    curr = gt_row + 1

    # ── Summary section ─────────────────────────────────────────────
    summ_hdr_row  = curr + 2    # 2 blank rows gap
    summ_data_row = summ_hdr_row + 1

    summ_hdrs = ['Department', 'Total Gross Pay', 'Total Net Pay',
                 'Total Employer Taxes', 'Total Payroll Cost']
    for c, h in enumerate(summ_hdrs, 1):
        write_cell(ws, summ_hdr_row, c, h, dark_fill, hdr_font, align='center')

    # Map summary cols to data cols: B→D, C→E, D→F, E→G
    data_col_map = {'B': 'D', 'C': 'E', 'D': 'F', 'E': 'G'}

    dept_summ_rows = {}
    sr = summ_data_row
    for dept in DEPT_ORDER:
        if dept not in dept_row_ranges:
            continue
        d_start, d_end = dept_row_ranges[dept]
        fill = cell_fill(DEPT_COLORS[dept])
        font = Font(name=FONT, size=10)

        write_cell(ws, sr, 1, dept, fill, Font(name=FONT, bold=True, size=10), align='left')
        for sum_col, data_col in data_col_map.items():
            c_idx = ord(sum_col) - 64
            write_cell(ws, sr, c_idx,
                       f'=SUM({data_col}{d_start}:{data_col}{d_end})',
                       fill, font, fmt=money_fmt, align='right')

        dept_summ_rows[dept] = sr
        sr += 1

    # Summary Total row
    total_row = sr
    for c in range(1, 6):
        write_cell(ws, total_row, c, None, grey_fill, gt_font,
                   align='left' if c == 1 else 'right')
    ws.cell(row=total_row, column=1).value = 'Total'
    for c, col in enumerate(['B', 'C', 'D', 'E'], 2):
        ws.cell(row=total_row, column=c).value = \
            f'=SUM({col}{summ_data_row}:{col}{total_row - 1})'
        ws.cell(row=total_row, column=c).number_format = money_fmt

    sr += 1

    # ── Check / Variance Analysis ────────────────────────────────────
    check_title_row = sr + 2
    check_hdr_row   = check_title_row + 1
    check_data_row  = check_hdr_row + 1

    title_font = Font(name=FONT, bold=True, size=11, color='FFFFFF')
    for c in range(1, 6):
        write_cell(ws, check_title_row, c, None, dark_fill, title_font, align='center')
    ws.cell(row=check_title_row, column=1).value = 'CHECK / VARIANCE ANALYSIS'
    ws.row_dimensions[check_title_row].height = 22

    chk_hdrs = ['Metric', 'Employee Grand Total', 'Dept Summary Total', 'Variance', 'Status']
    hdr2_font = Font(name=FONT, bold=True, size=10, color='FFFFFF')
    for c, h in enumerate(chk_hdrs, 1):
        write_cell(ws, check_hdr_row, c, h,
                   cell_fill('404040'), hdr2_font, align='center')
    ws.row_dimensions[check_hdr_row].height = 20

    # (label, employee-data-col, dept-summary-col-in-summary-section)
    checks = [
        ('Gross Pay',          'D', 'B'),
        ('Net Pay',            'E', 'C'),
        ('Employer Taxes',     'F', 'D'),
        ('Total Payroll Cost', 'G', 'E'),
    ]
    row_fill   = cell_fill('F2F2F2')
    green_fill = cell_fill('C6EFCE')
    row_font2  = Font(name=FONT, size=10)
    green_font = Font(name=FONT, bold=True, size=10, color='276221')
    var_fmt    = '$#,##0.00;[Red]($#,##0.00);"-"'

    for i, (label, ec, dc) in enumerate(checks):
        r = check_data_row + i
        write_cell(ws, r, 1, label,              row_fill,   Font(name=FONT, bold=True, size=10))
        write_cell(ws, r, 2, f'={ec}{gt_row}',   row_fill,   row_font2, fmt=money_fmt, align='right')
        write_cell(ws, r, 3, f'={dc}{total_row}', row_fill,  row_font2, fmt=money_fmt, align='right')
        write_cell(ws, r, 4, f'=B{r}-C{r}',      row_fill,   row_font2, fmt=var_fmt,   align='right')
        write_cell(ws, r, 5,
                   f'=IF(ABS(D{r})<0.005,"✓ Matched – No Variance","✗ VARIANCE DETECTED")',
                   green_fill, green_font, align='center')
        ws.row_dimensions[r].height = 18

    # ── Column widths ─────────────────────────────────────────────────
    col_widths = {'A': 30, 'B': 14, 'C': 26, 'D': 18, 'E': 16, 'F': 32, 'G': 18}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    wb.save(output_path)
    return dept_row_ranges, no_group_placeholder


# Placeholder so build_excel can return no_group; real flow passes it from outside
no_group_placeholder = []


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Payroll Report Generator')
    parser.add_argument('--input',  required=True, help='Path to Gusto payroll CSV')
    parser.add_argument('--output', required=True, help='Output Excel file path')
    parser.add_argument('--depts',  default=None,  help='Path to departments.json')
    args = parser.parse_args()

    # Default depts path: references/departments.json next to this script
    if args.depts is None:
        script_dir = Path(__file__).parent
        args.depts = str(script_dir.parent / 'references' / 'departments.json')

    print(f"Loading departments from: {args.depts}")
    lookup = build_lookup(args.depts)

    print(f"Reading payroll file: {args.input}")
    employees = read_payroll_file(args.input)
    print(f"  Found {len(employees)} employees.")

    employees, no_group = assign_and_sort(employees, lookup)

    if no_group:
        print(f"\n⚠️  UNMATCHED EMPLOYEES (assigned to 'No Group'):")
        for name in no_group:
            print(f"   - {name}")
    else:
        print("  All employees matched to a department.")

    print(f"\nBuilding Excel report: {args.output}")

    # Rebuild build_excel to accept no_group properly
    wb = Workbook()
    ws = wb.active
    ws.title = 'Payroll Report'

    dark_fill = cell_fill('595959')
    grey_fill = cell_fill(GREY)
    money_fmt = '$#,##0.00'

    hdr_cols = [
        'Name', 'Pay Date', 'Time Period',
        'Gross Pay – Total', 'Net Pay',
        'Employer Taxes & Contributions – Total', 'Total Payroll Cost',
    ]
    hdr_font = Font(name=FONT, bold=True, size=10, color='FFFFFF')
    for c, h in enumerate(hdr_cols, 1):
        write_cell(ws, 1, c, h, dark_fill, hdr_font, align='center')
    ws.row_dimensions[1].height = 30

    curr = 2
    dept_row_ranges = {}

    for emp in employees:
        dept  = emp['dept']
        color = DEPT_COLORS.get(dept, 'FFFFFF')
        fill  = cell_fill(color)
        font  = Font(name=FONT, size=10)

        if dept not in dept_row_ranges:
            dept_row_ranges[dept] = [curr, curr]
        else:
            dept_row_ranges[dept][1] = curr

        values = [
            emp['name'], emp['pay_date'], emp['period'],
            emp['gross'], emp['net'], emp['emp_tax'], emp['total'],
        ]
        for c, v in enumerate(values, 1):
            write_cell(ws, curr, c, v, fill, font,
                       fmt=money_fmt if c >= 4 else None,
                       align='left' if c <= 3 else 'right')
        curr += 1

    last_data_row = curr - 1
    gt_row  = curr
    gt_font = Font(name=FONT, bold=True, size=10)

    write_cell(ws, gt_row, 1, 'Grand Total', grey_fill, gt_font, align='left')
    write_cell(ws, gt_row, 2, None, grey_fill, gt_font)
    write_cell(ws, gt_row, 3, None, grey_fill, gt_font)
    for c, col in enumerate(['D', 'E', 'F', 'G'], 4):
        write_cell(ws, gt_row, c,
                   f'=SUM({col}2:{col}{last_data_row})',
                   grey_fill, gt_font, fmt=money_fmt, align='right')

    curr = gt_row + 1
    summ_hdr_row  = curr + 2
    summ_data_row = summ_hdr_row + 1

    summ_hdrs = ['Department', 'Total Gross Pay', 'Total Net Pay',
                 'Total Employer Taxes', 'Total Payroll Cost']
    for c, h in enumerate(summ_hdrs, 1):
        write_cell(ws, summ_hdr_row, c, h, dark_fill, hdr_font, align='center')

    data_col_map = {'B': 'D', 'C': 'E', 'D': 'F', 'E': 'G'}
    sr = summ_data_row
    for dept in DEPT_ORDER:
        if dept not in dept_row_ranges:
            continue
        d_start, d_end = dept_row_ranges[dept]
        fill = cell_fill(DEPT_COLORS[dept])
        font = Font(name=FONT, size=10)
        write_cell(ws, sr, 1, dept, fill, Font(name=FONT, bold=True, size=10), align='left')
        for sum_col, data_col in data_col_map.items():
            c_idx = ord(sum_col) - 64
            write_cell(ws, sr, c_idx,
                       f'=SUM({data_col}{d_start}:{data_col}{d_end})',
                       fill, font, fmt=money_fmt, align='right')
        sr += 1

    total_row = sr
    for c in range(1, 6):
        write_cell(ws, total_row, c, None, grey_fill, gt_font,
                   align='left' if c == 1 else 'right')
    ws.cell(row=total_row, column=1).value = 'Total'
    for c, col in enumerate(['B', 'C', 'D', 'E'], 2):
        ws.cell(row=total_row, column=c).value = \
            f'=SUM({col}{summ_data_row}:{col}{total_row - 1})'
        ws.cell(row=total_row, column=c).number_format = money_fmt

    sr += 1
    check_title_row = sr + 2
    check_hdr_row   = check_title_row + 1
    check_data_row  = check_hdr_row + 1

    title_font = Font(name=FONT, bold=True, size=11, color='FFFFFF')
    for c in range(1, 6):
        write_cell(ws, check_title_row, c, None, dark_fill, title_font, align='center')
    ws.cell(row=check_title_row, column=1).value = 'CHECK / VARIANCE ANALYSIS'
    ws.row_dimensions[check_title_row].height = 22

    for c, h in enumerate(['Metric', 'Employee Grand Total', 'Dept Summary Total',
                            'Variance', 'Status'], 1):
        write_cell(ws, check_hdr_row, c, h, cell_fill('404040'),
                   Font(name=FONT, bold=True, size=10, color='FFFFFF'), align='center')
    ws.row_dimensions[check_hdr_row].height = 20

    checks = [
        ('Gross Pay',          'D', 'B'),
        ('Net Pay',            'E', 'C'),
        ('Employer Taxes',     'F', 'D'),
        ('Total Payroll Cost', 'G', 'E'),
    ]
    row_fill   = cell_fill('F2F2F2')
    green_fill = cell_fill('C6EFCE')
    row_font2  = Font(name=FONT, size=10)
    green_font = Font(name=FONT, bold=True, size=10, color='276221')
    var_fmt    = '$#,##0.00;[Red]($#,##0.00);"-"'

    for i, (label, ec, dc) in enumerate(checks):
        r = check_data_row + i
        write_cell(ws, r, 1, label,               row_fill,   Font(name=FONT, bold=True, size=10))
        write_cell(ws, r, 2, f'={ec}{gt_row}',    row_fill,   row_font2, fmt=money_fmt, align='right')
        write_cell(ws, r, 3, f'={dc}{total_row}', row_fill,   row_font2, fmt=money_fmt, align='right')
        write_cell(ws, r, 4, f'=B{r}-C{r}',       row_fill,   row_font2, fmt=var_fmt,   align='right')
        write_cell(ws, r, 5,
                   f'=IF(ABS(D{r})<0.005,"✓ Matched – No Variance","✗ VARIANCE DETECTED")',
                   green_fill, green_font, align='center')
        ws.row_dimensions[r].height = 18

    col_widths = {'A': 30, 'B': 14, 'C': 26, 'D': 18, 'E': 16, 'F': 32, 'G': 18}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    wb.save(args.output)
    print(f"\n✅ Report saved to: {args.output}")

    if no_group:
        print(f"\n⚠️  Review these unmatched employees and add them to departments.json:")
        for name in no_group:
            print(f"   - {name}")


if __name__ == '__main__':
    main()
